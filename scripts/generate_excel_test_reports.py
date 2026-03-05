from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Sequence

from openpyxl import Workbook


@dataclass(frozen=True)
class ReportSeed:
    filename: str
    customer: str
    author: str
    site: str
    equipment: str
    manager: str
    work_date: str
    work_time: str
    work_detail: str


SEEDS: Sequence[ReportSeed] = (
    ReportSeed(
        filename="excel_report_01_hanbit_ljx8200.xlsx",
        customer="한빛정밀",
        author="김서준",
        site="인천 3공장",
        equipment="LJ-X8200",
        manager="박민아",
        work_date="2026-01-09",
        work_time="09:00-11:30",
        work_detail="라인 프로파일 센서 초기 교정 및 노이즈 맵 재측정",
    ),
    ReportSeed(
        filename="excel_report_02_mirae_visionflex.xlsx",
        customer="미래오토메이션",
        author="이재민",
        site="평택 라인 B",
        equipment="VisionFlex-Cam-12",
        manager="정다은",
        work_date="2026-01-11",
        work_time="13:10-15:00",
        work_detail="카메라 노출값 재튜닝 및 광원 플리커 점검",
    ),
    ReportSeed(
        filename="excel_report_03_sungjin_pressline.xlsx",
        customer="성진테크",
        author="최윤호",
        site="창원 프레스동",
        equipment="PressLine-QA-07",
        manager="윤성희",
        work_date="2026-01-13",
        work_time="07:40-09:20",
        work_detail="프레스 금형 높이 편차 보정 및 경고 임계치 수정",
    ),
    ReportSeed(
        filename="excel_report_04_daeil_packaging.xlsx",
        customer="대일패키징",
        author="오하늘",
        site="김해 포장라인",
        equipment="Seal-Inspect-3000",
        manager="임정우",
        work_date="2026-01-15",
        work_time="20:00-22:10",
        work_detail="포장 실링 불량 검출 민감도 상향 및 테스트 샘플 검증",
    ),
    ReportSeed(
        filename="excel_report_05_seoil_semicon.xlsx",
        customer="서일반도체",
        author="한도윤",
        site="청주 FAB-2",
        equipment="Wafer-Edge-Scan-R2",
        manager="강세린",
        work_date="2026-01-18",
        work_time="10:30-12:30",
        work_detail="웨이퍼 엣지 스캔 오차 추적 및 기준파일 업데이트",
    ),
    ReportSeed(
        filename="excel_report_06_ujin_logis.xlsx",
        customer="우진로지스",
        author="문지훈",
        site="군포 물류센터",
        equipment="Sorter-LiDAR-9",
        manager="배유진",
        work_date="2026-01-21",
        work_time="14:00-17:10",
        work_detail="분류기 LiDAR 축 정렬 및 야간 모드 감도 조정",
    ),
    ReportSeed(
        filename="excel_report_07_pioneer_glass.xlsx",
        customer="Pioneer Glass",
        author="J. Park",
        site="Asan Line-4",
        equipment="EdgeCheck-2D",
        manager="Mina Choi",
        work_date="2026-01-23",
        work_time="08:50-10:15",
        work_detail="Edge chip detection threshold calibration and retest",
    ),
    ReportSeed(
        filename="excel_report_08_koryo_food.xlsx",
        customer="고려푸드",
        author="서지안",
        site="광주 HACCP 구역",
        equipment="Label-Verify-MX",
        manager="백현주",
        work_date="2026-01-26",
        work_time="16:20-18:00",
        work_detail="라벨 인쇄 불량 OCR 판독률 개선 및 작업자 교육",
    ),
    ReportSeed(
        filename="excel_report_09_nova_battery.xlsx",
        customer="Nova Battery",
        author="Alex Kim",
        site="Ulsan Cell Plant",
        equipment="Cell-Stack-Inspector",
        manager="D. Lee",
        work_date="2026-01-28",
        work_time="11:00-13:40",
        work_detail="Stack alignment drift analysis and alarm profile tuning",
    ),
    ReportSeed(
        filename="excel_report_10_taesan_machinery.xlsx",
        customer="태산기계",
        author="신예린",
        site="대전 가공센터",
        equipment="CNC-Surface-Check-11",
        manager="권도형",
        work_date="2026-01-30",
        work_time="06:30-09:00",
        work_detail="표면 결함 기준값 재정의 및 시범 배치 측정",
    ),
)


def _key_value_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "작업보고"
    ws.append(["항목", "값"])
    ws.append(["고객사", seed.customer])
    ws.append(["작성자", seed.author])
    ws.append(["작업장소", seed.site])
    ws.append(["대상설비", seed.equipment])
    ws.append(["고객사 담당자 이름", seed.manager])
    ws.append(["작업 일자", seed.work_date])
    ws.append(["작업 시간", seed.work_time])
    ws.append(["작업 내용", seed.work_detail])


def _table_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "DailyLog"
    ws.append(["고객사", "작성자", "작업장소", "대상설비", "고객사 담당자 이름", "작업 일자", "작업 시간", "작업 내용"])
    ws.append([
        seed.customer,
        seed.author,
        seed.site,
        seed.equipment,
        seed.manager,
        seed.work_date,
        seed.work_time,
        seed.work_detail,
    ])
    ws.append(["비고", "", "", "", "", "", "", "진동 센서 로그 동시 확인"])


def _multi_sheet_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws1 = workbook.active
    ws1.title = "Summary"
    ws1.append(["구분", "내용"])
    ws1.append(["고객사", seed.customer])
    ws1.append(["작성자", seed.author])
    ws1.append(["작업 내용", seed.work_detail])

    ws2 = workbook.create_sheet("작업기록")
    ws2.append(["항목", "값"])
    ws2.append(["작업장소", seed.site])
    ws2.append(["대상설비", seed.equipment])
    ws2.append(["고객사 담당자 이름", seed.manager])
    ws2.append(["작업 일자", seed.work_date])
    ws2.append(["작업 시간", seed.work_time])


def _freeform_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "현장메모"
    ws.append(["현장 보고서"])
    ws.append([f"고객사={seed.customer}"])
    ws.append([f"작성자={seed.author}"])
    ws.append([f"작업장소={seed.site}"])
    ws.append([f"대상설비={seed.equipment}"])
    ws.append([f"고객사 담당자 이름={seed.manager}"])
    ws.append([f"작업 일자={seed.work_date} / 작업 시간={seed.work_time}"])
    ws.append([f"작업 내용: {seed.work_detail}"])


def _checklist_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "Checklist"
    ws.append(["체크", "세부항목", "결과", "메모"])
    ws.append(["v", "고객사", seed.customer, ""])
    ws.append(["v", "작성자", seed.author, ""])
    ws.append(["v", "작업장소", seed.site, ""])
    ws.append(["v", "대상설비", seed.equipment, ""])
    ws.append(["v", "고객사 담당자 이름", seed.manager, ""])
    ws.append(["v", "작업 일자", seed.work_date, ""])
    ws.append(["v", "작업 시간", seed.work_time, ""])
    ws.append(["v", "작업 내용", seed.work_detail, "후속점검 필요"])


def _timeline_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "Timeline"
    ws.append(["시간", "작업"])
    ws.append(["Start", f"고객사 {seed.customer} / 작업장소 {seed.site}"])
    ws.append(["T+20m", f"대상설비 {seed.equipment} 상태 점검"])
    ws.append(["T+40m", f"고객사 담당자 이름 {seed.manager} 입회 확인"])
    ws.append(["T+70m", f"작업 내용 {seed.work_detail}"])
    ws.append(["Meta", f"작성자 {seed.author} / 작업 일자 {seed.work_date} / 작업 시간 {seed.work_time}"])


def _incident_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "Incident"
    ws.append(["필드", "값", "비고"])
    ws.append(["고객사", seed.customer, "긴급 호출"])
    ws.append(["작성자", seed.author, "야간대응"])
    ws.append(["작업장소", seed.site, "라인 정지 12분"])
    ws.append(["대상설비", seed.equipment, "알람 코드 E-17"])
    ws.append(["고객사 담당자 이름", seed.manager, "현장 승인"])
    ws.append(["작업 일자", seed.work_date, ""])
    ws.append(["작업 시간", seed.work_time, ""])
    ws.append(["작업 내용", seed.work_detail, "재발방지 조치서 송부"])


def _bilingual_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "Bilingual"
    ws.append(["Field", "Value"])
    ws.append(["Customer / 고객사", seed.customer])
    ws.append(["Author / 작성자", seed.author])
    ws.append(["Work Site / 작업장소", seed.site])
    ws.append(["Target Equipment / 대상설비", seed.equipment])
    ws.append(["Customer Manager / 고객사 담당자 이름", seed.manager])
    ws.append(["Work Date / 작업 일자", seed.work_date])
    ws.append(["Work Time / 작업 시간", seed.work_time])
    ws.append(["Work Detail / 작업 내용", seed.work_detail])


def _ops_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "운영기록"
    ws.append(["No", "라벨", "값"])
    ws.append([1, "고객사", seed.customer])
    ws.append([2, "작성자", seed.author])
    ws.append([3, "작업장소", seed.site])
    ws.append([4, "대상설비", seed.equipment])
    ws.append([5, "고객사 담당자 이름", seed.manager])
    ws.append([6, "작업 일자", seed.work_date])
    ws.append([7, "작업 시간", seed.work_time])
    ws.append([8, "작업 내용", seed.work_detail])


def _summary_plus_detail_template(seed: ReportSeed, workbook: Workbook) -> None:
    ws1 = workbook.active
    ws1.title = "요약"
    ws1.append(["요약", f"{seed.customer} 현장 {seed.equipment} 점검"])
    ws1.append(["작성자", seed.author])
    ws1.append(["작업 일자", seed.work_date])

    ws2 = workbook.create_sheet("상세")
    ws2.append(["항목", "값"])
    ws2.append(["고객사", seed.customer])
    ws2.append(["작업장소", seed.site])
    ws2.append(["대상설비", seed.equipment])
    ws2.append(["고객사 담당자 이름", seed.manager])
    ws2.append(["작업 시간", seed.work_time])
    ws2.append(["작업 내용", seed.work_detail])


TEMPLATES: Sequence[Callable[[ReportSeed, Workbook], None]] = (
    _key_value_template,
    _table_template,
    _multi_sheet_template,
    _freeform_template,
    _checklist_template,
    _timeline_template,
    _incident_template,
    _bilingual_template,
    _ops_template,
    _summary_plus_detail_template,
)


def generate_reports(output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    generated: list[Path] = []
    for index, seed in enumerate(SEEDS):
        workbook = Workbook()
        template = TEMPLATES[index % len(TEMPLATES)]
        template(seed, workbook)

        destination = output_dir / seed.filename
        workbook.save(destination)
        workbook.close()
        generated.append(destination)

    return generated


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate sample Excel maintenance reports.")
    parser.add_argument(
        "--output-dir",
        default="uploads/excel-test-reports",
        help="Directory for generated xlsx files",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    files = generate_reports(output_dir)

    print(f"Generated {len(files)} Excel reports in {output_dir}")
    for file_path in files:
        print(f"- {file_path}")


if __name__ == "__main__":
    main()
