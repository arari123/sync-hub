from __future__ import annotations

import hashlib
import json
import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

from .budget_logic import labor_budget_amount, normalize_phase, to_number

SUMMARY_SHEET = "요약"
MATERIAL_SHEET = "재료비"
LABOR_SHEET = "인건비"
EXPENSE_SHEET = "경비"
META_SHEET = "_meta"

DATA_START_ROW = 5
HEADER_ROW = 4

TEMPLATE_VERSION = "budget-excel.v1"
EXPORT_SCOPE = "execution_only"

SHEET_ORDER = [SUMMARY_SHEET, MATERIAL_SHEET, LABOR_SHEET, EXPENSE_SHEET, META_SHEET]

MATERIAL_HEADERS = [
    "No",
    "설비명",
    "단계",
    "유닛",
    "파츠",
    "스펙",
    "수량(예산)",
    "단가(예산)",
    "예산금액(수식)",
    "집행금액(입력)",
    "메모",
]

LABOR_HEADERS = [
    "No",
    "설비명",
    "단계",
    "구분",
    "업무명",
    "인력유형",
    "단위",
    "수량(예산)",
    "인원(예산)",
    "예산금액",
    "집행금액(입력)",
    "메모",
]

EXPENSE_HEADERS = [
    "No",
    "설비명",
    "단계",
    "구분",
    "항목명",
    "산출근거",
    "수량(예산)",
    "예산금액",
    "집행금액(입력)",
    "메모",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1E3A8A")
HEADER_FONT = Font(name="Malgun Gothic", bold=True, size=11, color="FFFFFF")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="0F172A")
TITLE_FONT = Font(name="Malgun Gothic", bold=True, size=14, color="FFFFFF")
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="E2E8F0")
TOTAL_FONT = Font(name="Malgun Gothic", bold=True, size=10, color="1F2937")
BODY_FONT = Font(name="Malgun Gothic", size=10, color="111827")

CURRENCY_FORMAT = '"₩"#,##0'
QUANTITY_FORMAT = "#,##0.00"
PERCENT_FORMAT = "0.0%"

PROTECTION_PASSWORD = "sync-hub"
MAX_ERRORS = 80


class BudgetExcelValidationError(Exception):
    def __init__(self, errors: list[str]):
        self.errors = [str(item) for item in (errors or []) if str(item).strip()][:MAX_ERRORS]
        message = "\n".join(self.errors) if self.errors else "엑셀 양식 검증에 실패했습니다."
        super().__init__(message)

    def to_detail(self) -> str:
        return "\n".join(self.errors) if self.errors else "엑셀 양식 검증에 실패했습니다."


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except Exception:  # noqa: BLE001
        return 0.0


def _normalize_phase_code(value: Any) -> str:
    raw = _normalize_text(value).lower()
    if raw in {"제작", "fabrication", "fab"}:
        return "fabrication"
    if raw in {"설치", "installation", "install"}:
        return "installation"
    return ""


def _phase_label(phase: str) -> str:
    return "설치" if normalize_phase(phase) == "installation" else "제작"


def _safe_formula(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = text.upper().replace(" ", "")
    if text.startswith("="):
        text = text[1:]
    # Normalize locale/editor differences while keeping formula intent strict.
    text = text.replace(";", ",")
    text = text.replace("$", "")
    text = text.replace("'", "")
    text = text.replace("_XLFN.", "")
    text = text.replace("@", "")
    text = re.sub(r"(?<=[,(])(-?\d+)\.0(?=[,)])", r"\1", text)
    return text


def _template_signature() -> str:
    payload = {
        "template_version": TEMPLATE_VERSION,
        "export_scope": EXPORT_SCOPE,
        "header_row": HEADER_ROW,
        "data_start_row": DATA_START_ROW,
        "material_headers": MATERIAL_HEADERS,
        "labor_headers": LABOR_HEADERS,
        "expense_headers": EXPENSE_HEADERS,
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def _material_unit_multiplier(settings: dict[str, Any], equipment: str, phase: str, unit_name: str) -> float:
    unit_counts = settings.get("material_unit_counts")
    if not isinstance(unit_counts, dict):
        return 1.0
    scope_key = f"{equipment}::{phase}::{unit_name}" if unit_name else ""
    if not scope_key:
        return 1.0
    count = _to_number(unit_counts.get(scope_key))
    if count <= 0:
        return 1.0
    return float(max(1, int(count)))


def _build_material_rows(detail_payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    settings = detail_payload.get("budget_settings") if isinstance(detail_payload.get("budget_settings"), dict) else {}

    execution_map: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for item in detail_payload.get("execution_material_items") or []:
        equipment = _normalize_text(item.get("equipment_name"))
        phase = normalize_phase(item.get("phase") or "fabrication")
        unit_name = _normalize_text(item.get("unit_name"))
        part_name = _normalize_text(item.get("part_name"))
        spec = _normalize_text(item.get("spec"))
        key = (equipment, phase, unit_name, part_name, spec)
        bucket = execution_map.setdefault(key, {"executed": 0.0, "memo": ""})
        bucket["executed"] += _to_number(item.get("executed_amount"))
        memo = _normalize_text(item.get("memo"))
        if memo and not bucket["memo"]:
            bucket["memo"] = memo

    for item in detail_payload.get("material_items") or []:
        equipment = _normalize_text(item.get("equipment_name")) or "미지정 설비"
        phase = normalize_phase(item.get("phase") or "fabrication")
        unit_name = _normalize_text(item.get("unit_name"))
        part_name = _normalize_text(item.get("part_name"))
        spec = _normalize_text(item.get("spec"))
        quantity = _to_number(item.get("quantity"))
        unit_price = _to_number(item.get("unit_price"))
        unit_scope_name = unit_name or part_name
        unit_multiplier = _material_unit_multiplier(settings, equipment, phase, unit_scope_name)
        budget_amount = quantity * unit_price * unit_multiplier

        key = (equipment, phase, unit_name, part_name, spec)
        execution = execution_map.pop(key, None) or {}

        rows.append(
            {
                "equipment_name": equipment,
                "phase": phase,
                "unit_name": unit_name,
                "part_name": part_name,
                "spec": spec,
                "quantity": quantity,
                "unit_price": unit_price,
                "budget_amount": budget_amount,
                "executed_amount": _to_number(execution.get("executed") if execution else item.get("executed_amount")),
                "memo": _normalize_text(execution.get("memo") if execution else item.get("memo")),
            }
        )

    for key, execution in execution_map.items():
        equipment, phase, unit_name, part_name, spec = key
        rows.append(
            {
                "equipment_name": equipment or "미지정 설비",
                "phase": phase,
                "unit_name": unit_name,
                "part_name": part_name,
                "spec": spec,
                "quantity": 0.0,
                "unit_price": 0.0,
                "budget_amount": 0.0,
                "executed_amount": _to_number(execution.get("executed")),
                "memo": _normalize_text(execution.get("memo")),
            }
        )

    rows.sort(
        key=lambda row: (
            _normalize_text(row.get("equipment_name")),
            normalize_phase(row.get("phase") or "fabrication"),
            _normalize_text(row.get("unit_name")),
            _normalize_text(row.get("part_name")),
            _normalize_text(row.get("spec")),
        )
    )
    return rows


def _build_labor_rows(detail_payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    settings = detail_payload.get("budget_settings") if isinstance(detail_payload.get("budget_settings"), dict) else {}

    execution_map: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for item in detail_payload.get("execution_labor_items") or []:
        equipment = _normalize_text(item.get("equipment_name"))
        phase = normalize_phase(item.get("phase") or "fabrication")
        staffing_type = "외주" if _normalize_text(item.get("staffing_type")) == "외주" else "자체"
        task_name = _normalize_text(item.get("task_name"))
        worker_type = _normalize_text(item.get("worker_type"))
        key = (equipment, phase, staffing_type, task_name, worker_type)
        bucket = execution_map.setdefault(key, {"executed": 0.0, "memo": ""})
        bucket["executed"] += _to_number(item.get("executed_amount"))
        memo = _normalize_text(item.get("memo"))
        if memo and not bucket["memo"]:
            bucket["memo"] = memo

    for item in detail_payload.get("labor_items") or []:
        equipment = _normalize_text(item.get("equipment_name")) or "미지정 설비"
        phase = normalize_phase(item.get("phase") or "fabrication")
        staffing_type = "외주" if _normalize_text(item.get("staffing_type")) == "외주" else "자체"
        task_name = _normalize_text(item.get("task_name"))
        worker_type = _normalize_text(item.get("worker_type"))
        unit = _normalize_text(item.get("unit")) or "H"
        quantity = _to_number(item.get("quantity"))
        headcount = _to_number(item.get("headcount")) or 1.0
        budget_amount = labor_budget_amount(item, settings=settings)

        key = (equipment, phase, staffing_type, task_name, worker_type)
        execution = execution_map.pop(key, None) or {}

        rows.append(
            {
                "equipment_name": equipment,
                "phase": phase,
                "staffing_type": staffing_type,
                "task_name": task_name,
                "worker_type": worker_type,
                "unit": unit,
                "quantity": quantity,
                "headcount": headcount,
                "budget_amount": budget_amount,
                "executed_amount": _to_number(execution.get("executed") if execution else item.get("executed_amount")),
                "memo": _normalize_text(execution.get("memo") if execution else item.get("memo")),
            }
        )

    for key, execution in execution_map.items():
        equipment, phase, staffing_type, task_name, worker_type = key
        rows.append(
            {
                "equipment_name": equipment or "미지정 설비",
                "phase": phase,
                "staffing_type": staffing_type,
                "task_name": task_name,
                "worker_type": worker_type,
                "unit": "H",
                "quantity": 0.0,
                "headcount": 1.0,
                "budget_amount": 0.0,
                "executed_amount": _to_number(execution.get("executed")),
                "memo": _normalize_text(execution.get("memo")),
            }
        )

    rows.sort(
        key=lambda row: (
            _normalize_text(row.get("equipment_name")),
            normalize_phase(row.get("phase") or "fabrication"),
            _normalize_text(row.get("staffing_type")),
            _normalize_text(row.get("task_name")),
            _normalize_text(row.get("worker_type")),
        )
    )
    return rows


def _build_expense_rows(detail_payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    execution_map: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for item in detail_payload.get("execution_expense_items") or []:
        equipment = _normalize_text(item.get("equipment_name"))
        phase = normalize_phase(item.get("phase") or "fabrication")
        expense_type = "외주" if _normalize_text(item.get("expense_type")) == "외주" else "자체"
        expense_name = _normalize_text(item.get("expense_name"))
        basis = _normalize_text(item.get("basis"))
        key = (equipment, phase, expense_type, expense_name, basis)
        bucket = execution_map.setdefault(key, {"executed": 0.0, "memo": ""})
        bucket["executed"] += _to_number(item.get("executed_amount"))
        memo = _normalize_text(item.get("memo"))
        if memo and not bucket["memo"]:
            bucket["memo"] = memo

    for item in detail_payload.get("expense_items") or []:
        equipment = _normalize_text(item.get("equipment_name")) or "미지정 설비"
        phase = normalize_phase(item.get("phase") or "fabrication")
        expense_type = "외주" if _normalize_text(item.get("expense_type")) == "외주" else "자체"
        expense_name = _normalize_text(item.get("expense_name"))
        basis = _normalize_text(item.get("basis"))
        quantity = _to_number(item.get("quantity"))
        budget_amount = _to_number(item.get("amount"))

        key = (equipment, phase, expense_type, expense_name, basis)
        execution = execution_map.pop(key, None) or {}

        rows.append(
            {
                "equipment_name": equipment,
                "phase": phase,
                "expense_type": expense_type,
                "expense_name": expense_name,
                "basis": basis,
                "quantity": quantity,
                "budget_amount": budget_amount,
                "executed_amount": _to_number(execution.get("executed") if execution else item.get("executed_amount")),
                "memo": _normalize_text(execution.get("memo") if execution else item.get("memo")),
            }
        )

    for key, execution in execution_map.items():
        equipment, phase, expense_type, expense_name, basis = key
        rows.append(
            {
                "equipment_name": equipment or "미지정 설비",
                "phase": phase,
                "expense_type": expense_type,
                "expense_name": expense_name,
                "basis": basis,
                "quantity": 0.0,
                "budget_amount": 0.0,
                "executed_amount": _to_number(execution.get("executed")),
                "memo": _normalize_text(execution.get("memo")),
            }
        )

    rows.sort(
        key=lambda row: (
            _normalize_text(row.get("equipment_name")),
            normalize_phase(row.get("phase") or "fabrication"),
            _normalize_text(row.get("expense_type")),
            _normalize_text(row.get("expense_name")),
            _normalize_text(row.get("basis")),
        )
    )
    return rows


def _style_title_row(ws, title: str, max_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def _style_header_row(ws, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[HEADER_ROW].height = 22


def _apply_body_style(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if row % 2 == 0:
            cell.fill = ALT_ROW_FILL


def _protect_sheet(ws) -> None:
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD


def _write_material_sheet(wb: Workbook, rows: list[dict[str, Any]], row_count: int) -> None:
    ws = wb.create_sheet(MATERIAL_SHEET)
    _style_title_row(ws, "재료비 집행 입력 시트", len(MATERIAL_HEADERS))
    _style_header_row(ws, MATERIAL_HEADERS)

    row_end = DATA_START_ROW + row_count - 1
    ws["A2"] = "총 예산"
    ws["B2"] = f"=SUM(I{DATA_START_ROW}:I{row_end})"
    ws["C2"] = "총 집행"
    ws["D2"] = f"=SUM(J{DATA_START_ROW}:J{row_end})"

    for cell_name in ("A2", "C2"):
        ws[cell_name].font = TOTAL_FONT
        ws[cell_name].fill = TOTAL_FILL
        ws[cell_name].alignment = Alignment(horizontal="left", vertical="center")
    for cell_name in ("B2", "D2"):
        ws[cell_name].font = TOTAL_FONT
        ws[cell_name].fill = TOTAL_FILL
        ws[cell_name].number_format = CURRENCY_FORMAT
        ws[cell_name].alignment = Alignment(horizontal="right", vertical="center")

    for offset in range(row_count):
        row_idx = DATA_START_ROW + offset
        source = rows[offset] if offset < len(rows) else {}
        ws.cell(row=row_idx, column=1, value=offset + 1)
        ws.cell(row=row_idx, column=2, value=_normalize_text(source.get("equipment_name")))
        ws.cell(row=row_idx, column=3, value=_phase_label(source.get("phase") or "fabrication"))
        ws.cell(row=row_idx, column=4, value=_normalize_text(source.get("unit_name")))
        ws.cell(row=row_idx, column=5, value=_normalize_text(source.get("part_name")))
        ws.cell(row=row_idx, column=6, value=_normalize_text(source.get("spec")))

        qty_cell = ws.cell(row=row_idx, column=7, value=_to_number(source.get("quantity")))
        price_cell = ws.cell(row=row_idx, column=8, value=_to_number(source.get("unit_price")))
        budget_formula_cell = ws.cell(row=row_idx, column=9, value=f"=IFERROR(G{row_idx}*H{row_idx},0)")
        execution_cell = ws.cell(row=row_idx, column=10, value=_to_number(source.get("executed_amount")))
        memo_cell = ws.cell(row=row_idx, column=11, value=_normalize_text(source.get("memo")))

        qty_cell.number_format = QUANTITY_FORMAT
        price_cell.number_format = CURRENCY_FORMAT
        budget_formula_cell.number_format = CURRENCY_FORMAT
        execution_cell.number_format = CURRENCY_FORMAT
        memo_cell.alignment = Alignment(horizontal="left", vertical="center")

        _apply_body_style(ws, row_idx, len(MATERIAL_HEADERS))

        execution_cell.protection = Protection(locked=False)
        memo_cell.protection = Protection(locked=False)

    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:K{row_end}"

    widths = [6, 22, 10, 18, 18, 18, 12, 14, 16, 16, 22]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    _protect_sheet(ws)


def _write_labor_sheet(wb: Workbook, rows: list[dict[str, Any]], row_count: int) -> None:
    ws = wb.create_sheet(LABOR_SHEET)
    _style_title_row(ws, "인건비 집행 입력 시트", len(LABOR_HEADERS))
    _style_header_row(ws, LABOR_HEADERS)

    row_end = DATA_START_ROW + row_count - 1
    ws["A2"] = "총 예산"
    ws["B2"] = f"=SUM(J{DATA_START_ROW}:J{row_end})"
    ws["C2"] = "총 집행"
    ws["D2"] = f"=SUM(K{DATA_START_ROW}:K{row_end})"

    for cell_name in ("A2", "C2"):
        ws[cell_name].font = TOTAL_FONT
        ws[cell_name].fill = TOTAL_FILL
        ws[cell_name].alignment = Alignment(horizontal="left", vertical="center")
    for cell_name in ("B2", "D2"):
        ws[cell_name].font = TOTAL_FONT
        ws[cell_name].fill = TOTAL_FILL
        ws[cell_name].number_format = CURRENCY_FORMAT
        ws[cell_name].alignment = Alignment(horizontal="right", vertical="center")

    for offset in range(row_count):
        row_idx = DATA_START_ROW + offset
        source = rows[offset] if offset < len(rows) else {}

        ws.cell(row=row_idx, column=1, value=offset + 1)
        ws.cell(row=row_idx, column=2, value=_normalize_text(source.get("equipment_name")))
        ws.cell(row=row_idx, column=3, value=_phase_label(source.get("phase") or "fabrication"))
        ws.cell(row=row_idx, column=4, value=_normalize_text(source.get("staffing_type")) or "자체")
        ws.cell(row=row_idx, column=5, value=_normalize_text(source.get("task_name")))
        ws.cell(row=row_idx, column=6, value=_normalize_text(source.get("worker_type")))
        ws.cell(row=row_idx, column=7, value=_normalize_text(source.get("unit")) or "H")

        quantity_cell = ws.cell(row=row_idx, column=8, value=_to_number(source.get("quantity")))
        headcount_cell = ws.cell(row=row_idx, column=9, value=_to_number(source.get("headcount")) or 1.0)
        budget_cell = ws.cell(row=row_idx, column=10, value=_to_number(source.get("budget_amount")))
        execution_cell = ws.cell(row=row_idx, column=11, value=_to_number(source.get("executed_amount")))
        memo_cell = ws.cell(row=row_idx, column=12, value=_normalize_text(source.get("memo")))

        quantity_cell.number_format = QUANTITY_FORMAT
        headcount_cell.number_format = QUANTITY_FORMAT
        budget_cell.number_format = CURRENCY_FORMAT
        execution_cell.number_format = CURRENCY_FORMAT
        memo_cell.alignment = Alignment(horizontal="left", vertical="center")

        _apply_body_style(ws, row_idx, len(LABOR_HEADERS))
        execution_cell.protection = Protection(locked=False)
        memo_cell.protection = Protection(locked=False)

    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:L{row_end}"

    widths = [6, 22, 10, 10, 18, 14, 9, 12, 10, 16, 16, 22]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    _protect_sheet(ws)


def _write_expense_sheet(wb: Workbook, rows: list[dict[str, Any]], row_count: int) -> None:
    ws = wb.create_sheet(EXPENSE_SHEET)
    _style_title_row(ws, "경비 집행 입력 시트", len(EXPENSE_HEADERS))
    _style_header_row(ws, EXPENSE_HEADERS)

    row_end = DATA_START_ROW + row_count - 1
    ws["A2"] = "총 예산"
    ws["B2"] = f"=SUM(H{DATA_START_ROW}:H{row_end})"
    ws["C2"] = "총 집행"
    ws["D2"] = f"=SUM(I{DATA_START_ROW}:I{row_end})"

    for cell_name in ("A2", "C2"):
        ws[cell_name].font = TOTAL_FONT
        ws[cell_name].fill = TOTAL_FILL
        ws[cell_name].alignment = Alignment(horizontal="left", vertical="center")
    for cell_name in ("B2", "D2"):
        ws[cell_name].font = TOTAL_FONT
        ws[cell_name].fill = TOTAL_FILL
        ws[cell_name].number_format = CURRENCY_FORMAT
        ws[cell_name].alignment = Alignment(horizontal="right", vertical="center")

    for offset in range(row_count):
        row_idx = DATA_START_ROW + offset
        source = rows[offset] if offset < len(rows) else {}

        ws.cell(row=row_idx, column=1, value=offset + 1)
        ws.cell(row=row_idx, column=2, value=_normalize_text(source.get("equipment_name")))
        ws.cell(row=row_idx, column=3, value=_phase_label(source.get("phase") or "fabrication"))
        ws.cell(row=row_idx, column=4, value=_normalize_text(source.get("expense_type")) or "자체")
        ws.cell(row=row_idx, column=5, value=_normalize_text(source.get("expense_name")))
        ws.cell(row=row_idx, column=6, value=_normalize_text(source.get("basis")))

        quantity_cell = ws.cell(row=row_idx, column=7, value=_to_number(source.get("quantity")))
        budget_cell = ws.cell(row=row_idx, column=8, value=_to_number(source.get("budget_amount")))
        execution_cell = ws.cell(row=row_idx, column=9, value=_to_number(source.get("executed_amount")))
        memo_cell = ws.cell(row=row_idx, column=10, value=_normalize_text(source.get("memo")))

        quantity_cell.number_format = QUANTITY_FORMAT
        budget_cell.number_format = CURRENCY_FORMAT
        execution_cell.number_format = CURRENCY_FORMAT
        memo_cell.alignment = Alignment(horizontal="left", vertical="center")

        _apply_body_style(ws, row_idx, len(EXPENSE_HEADERS))
        execution_cell.protection = Protection(locked=False)
        memo_cell.protection = Protection(locked=False)

    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:J{row_end}"

    widths = [6, 22, 10, 10, 18, 20, 12, 16, 16, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    _protect_sheet(ws)


def _write_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(SUMMARY_SHEET)
    _style_title_row(ws, "예산 종합 요약", 7)

    headers = ["구분", "예산", "집행", "잔액", "집행률"]
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        (5, "재료비", f"='{MATERIAL_SHEET}'!B2", f"='{MATERIAL_SHEET}'!D2"),
        (6, "인건비", f"='{LABOR_SHEET}'!B2", f"='{LABOR_SHEET}'!D2"),
        (7, "경비", f"='{EXPENSE_SHEET}'!B2", f"='{EXPENSE_SHEET}'!D2"),
        (8, "총계", "=SUM(B5:B7)", "=SUM(C5:C7)"),
    ]

    for row_idx, label, budget_formula, execution_formula in rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=budget_formula)
        ws.cell(row=row_idx, column=3, value=execution_formula)
        ws.cell(row=row_idx, column=4, value=f"=B{row_idx}-C{row_idx}")
        ws.cell(row=row_idx, column=5, value=f"=IFERROR(C{row_idx}/B{row_idx},0)")

        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = BODY_FONT if row_idx < 8 else TOTAL_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            if row_idx == 8:
                cell.fill = TOTAL_FILL

        for col in (2, 3, 4):
            ws.cell(row=row_idx, column=col).number_format = CURRENCY_FORMAT
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=5).number_format = PERCENT_FORMAT
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:E8"
    widths = [14, 16, 16, 16, 12]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    _protect_sheet(ws)


def _write_meta_sheet(
    wb: Workbook,
    *,
    project_id: int,
    version_id: int,
    material_row_count: int,
    labor_row_count: int,
    expense_row_count: int,
) -> None:
    ws = wb.create_sheet(META_SHEET)
    ws["A1"] = "template_version"
    ws["B1"] = TEMPLATE_VERSION
    ws["A2"] = "template_signature"
    ws["B2"] = _template_signature()
    ws["A3"] = "export_scope"
    ws["B3"] = EXPORT_SCOPE

    ws["A4"] = "material_row_count"
    ws["B4"] = material_row_count
    ws["A5"] = "labor_row_count"
    ws["B5"] = labor_row_count
    ws["A6"] = "expense_row_count"
    ws["B6"] = expense_row_count

    ws["A7"] = "data_start_row"
    ws["B7"] = DATA_START_ROW
    ws["A8"] = "project_id"
    ws["B8"] = int(project_id)
    ws["A9"] = "version_id"
    ws["B9"] = int(version_id)

    ws.sheet_state = "hidden"


def build_budget_excel_bytes(project: Any, version: Any, detail_payload: dict[str, Any]) -> bytes:
    material_rows = _build_material_rows(detail_payload)
    labor_rows = _build_labor_rows(detail_payload)
    expense_rows = _build_expense_rows(detail_payload)

    material_row_count = max(1, len(material_rows))
    labor_row_count = max(1, len(labor_rows))
    expense_row_count = max(1, len(expense_rows))

    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb)
    _write_material_sheet(wb, material_rows, material_row_count)
    _write_labor_sheet(wb, labor_rows, labor_row_count)
    _write_expense_sheet(wb, expense_rows, expense_row_count)
    _write_meta_sheet(
        wb,
        project_id=int(getattr(project, "id", 0) or 0),
        version_id=int(getattr(version, "id", 0) or 0),
        material_row_count=material_row_count,
        labor_row_count=labor_row_count,
        expense_row_count=expense_row_count,
    )

    # Force sheet ordering for deterministic signature checks.
    wb._sheets = [wb[name] for name in SHEET_ORDER]  # noqa: SLF001

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _read_meta(ws) -> dict[str, Any]:
    return {
        "template_version": ws["B1"].value,
        "template_signature": ws["B2"].value,
        "export_scope": ws["B3"].value,
        "material_row_count": ws["B4"].value,
        "labor_row_count": ws["B5"].value,
        "expense_row_count": ws["B6"].value,
        "data_start_row": ws["B7"].value,
    }


def _add_error(errors: list[str], message: str) -> None:
    if len(errors) >= MAX_ERRORS:
        return
    errors.append(message)


def _expect_headers(ws, headers: list[str], errors: list[str]) -> None:
    for idx, expected in enumerate(headers, start=1):
        address = f"{get_column_letter(idx)}{HEADER_ROW}"
        actual = _normalize_text(ws[address].value)
        if actual != expected:
            _add_error(errors, f"{ws.title}!{address}: 헤더가 변경되었습니다. 기대값='{expected}', 현재값='{actual}'")


def _expect_sheet_protection(ws, errors: list[str]) -> None:
    if not ws.protection.sheet:
        _add_error(errors, f"{ws.title}: 시트 보호가 해제되었습니다.")


def _expect_locked(cell, ws_name: str, errors: list[str]) -> None:
    if not bool(cell.protection.locked):
        _add_error(errors, f"{ws_name}!{cell.coordinate}: 보호 셀이 잠금 해제되었습니다.")


def _expect_unlocked(cell, ws_name: str, errors: list[str]) -> None:
    if bool(cell.protection.locked):
        _add_error(errors, f"{ws_name}!{cell.coordinate}: 입력 셀이 잠금 상태입니다.")


def _validate_material_sheet(ws, row_count: int, errors: list[str]) -> None:
    _expect_sheet_protection(ws, errors)
    _expect_headers(ws, MATERIAL_HEADERS, errors)

    row_end = DATA_START_ROW + row_count - 1
    expected_budget_formula = _safe_formula(f"=SUM(I{DATA_START_ROW}:I{row_end})")
    expected_exec_formula = _safe_formula(f"=SUM(J{DATA_START_ROW}:J{row_end})")
    if _safe_formula(ws["B2"].value) != expected_budget_formula:
        _add_error(errors, "재료비!B2: 총 예산 수식이 변경되었습니다.")
    if _safe_formula(ws["D2"].value) != expected_exec_formula:
        _add_error(errors, "재료비!D2: 총 집행 수식이 변경되었습니다.")
    _expect_locked(ws["B2"], ws.title, errors)
    _expect_locked(ws["D2"], ws.title, errors)

    for row_idx in range(DATA_START_ROW, row_end + 1):
        formula_cell = ws[f"I{row_idx}"]
        expected_formula = _safe_formula(f"=IFERROR(G{row_idx}*H{row_idx},0)")
        if _safe_formula(formula_cell.value) != expected_formula:
            _add_error(errors, f"재료비!I{row_idx}: 예산금액 수식이 변경되었습니다.")
        _expect_locked(formula_cell, ws.title, errors)
        _expect_unlocked(ws[f"J{row_idx}"], ws.title, errors)
        _expect_unlocked(ws[f"K{row_idx}"], ws.title, errors)


def _validate_labor_sheet(ws, row_count: int, errors: list[str]) -> None:
    _expect_sheet_protection(ws, errors)
    _expect_headers(ws, LABOR_HEADERS, errors)

    row_end = DATA_START_ROW + row_count - 1
    expected_budget_formula = _safe_formula(f"=SUM(J{DATA_START_ROW}:J{row_end})")
    expected_exec_formula = _safe_formula(f"=SUM(K{DATA_START_ROW}:K{row_end})")
    if _safe_formula(ws["B2"].value) != expected_budget_formula:
        _add_error(errors, "인건비!B2: 총 예산 수식이 변경되었습니다.")
    if _safe_formula(ws["D2"].value) != expected_exec_formula:
        _add_error(errors, "인건비!D2: 총 집행 수식이 변경되었습니다.")
    _expect_locked(ws["B2"], ws.title, errors)
    _expect_locked(ws["D2"], ws.title, errors)

    for row_idx in range(DATA_START_ROW, row_end + 1):
        _expect_unlocked(ws[f"K{row_idx}"], ws.title, errors)
        _expect_unlocked(ws[f"L{row_idx}"], ws.title, errors)


def _validate_expense_sheet(ws, row_count: int, errors: list[str]) -> None:
    _expect_sheet_protection(ws, errors)
    _expect_headers(ws, EXPENSE_HEADERS, errors)

    row_end = DATA_START_ROW + row_count - 1
    expected_budget_formula = _safe_formula(f"=SUM(H{DATA_START_ROW}:H{row_end})")
    expected_exec_formula = _safe_formula(f"=SUM(I{DATA_START_ROW}:I{row_end})")
    if _safe_formula(ws["B2"].value) != expected_budget_formula:
        _add_error(errors, "경비!B2: 총 예산 수식이 변경되었습니다.")
    if _safe_formula(ws["D2"].value) != expected_exec_formula:
        _add_error(errors, "경비!D2: 총 집행 수식이 변경되었습니다.")
    _expect_locked(ws["B2"], ws.title, errors)
    _expect_locked(ws["D2"], ws.title, errors)

    for row_idx in range(DATA_START_ROW, row_end + 1):
        _expect_unlocked(ws[f"I{row_idx}"], ws.title, errors)
        _expect_unlocked(ws[f"J{row_idx}"], ws.title, errors)


def _validate_summary_sheet(ws, errors: list[str]) -> None:
    _expect_sheet_protection(ws, errors)

    expected = {
        "B5": f"='{MATERIAL_SHEET}'!B2",
        "C5": f"='{MATERIAL_SHEET}'!D2",
        "B6": f"='{LABOR_SHEET}'!B2",
        "C6": f"='{LABOR_SHEET}'!D2",
        "B7": f"='{EXPENSE_SHEET}'!B2",
        "C7": f"='{EXPENSE_SHEET}'!D2",
        "B8": "=SUM(B5:B7)",
        "C8": "=SUM(C5:C7)",
        "D5": "=B5-C5",
        "D6": "=B6-C6",
        "D7": "=B7-C7",
        "D8": "=B8-C8",
        "E5": "=IFERROR(C5/B5,0)",
        "E6": "=IFERROR(C6/B6,0)",
        "E7": "=IFERROR(C7/B7,0)",
        "E8": "=IFERROR(C8/B8,0)",
    }
    for address, formula in expected.items():
        if _safe_formula(ws[address].value) != _safe_formula(formula):
            _add_error(errors, f"요약!{address}: 요약 수식이 변경되었습니다.")
        _expect_locked(ws[address], ws.title, errors)


def _validate_template(workbook) -> dict[str, int]:
    errors: list[str] = []

    if workbook.sheetnames != SHEET_ORDER:
        _add_error(errors, f"시트 구성이 변경되었습니다. 기대={SHEET_ORDER}, 현재={workbook.sheetnames}")

    if META_SHEET not in workbook.sheetnames:
        _add_error(errors, "_meta 시트가 없습니다.")
        raise BudgetExcelValidationError(errors)

    meta = _read_meta(workbook[META_SHEET])
    if _normalize_text(meta.get("template_version")) != TEMPLATE_VERSION:
        _add_error(errors, "_meta!B1: 템플릿 버전이 일치하지 않습니다.")
    if _normalize_text(meta.get("template_signature")) != _template_signature():
        _add_error(errors, "_meta!B2: 템플릿 서명이 일치하지 않습니다.")
    if _normalize_text(meta.get("export_scope")) != EXPORT_SCOPE:
        _add_error(errors, "_meta!B3: 내보내기 범위 값이 일치하지 않습니다.")

    data_start_row = int(_to_number(meta.get("data_start_row")) or 0)
    if data_start_row != DATA_START_ROW:
        _add_error(errors, f"_meta!B7: 데이터 시작 행이 변경되었습니다. 기대={DATA_START_ROW}, 현재={data_start_row}")

    row_counts = {
        MATERIAL_SHEET: int(_to_number(meta.get("material_row_count")) or 0),
        LABOR_SHEET: int(_to_number(meta.get("labor_row_count")) or 0),
        EXPENSE_SHEET: int(_to_number(meta.get("expense_row_count")) or 0),
    }
    for sheet_name, count in row_counts.items():
        if count <= 0:
            _add_error(errors, f"_meta: {sheet_name} 행 수 정보가 유효하지 않습니다.")

    if SUMMARY_SHEET in workbook.sheetnames:
        _validate_summary_sheet(workbook[SUMMARY_SHEET], errors)
    if MATERIAL_SHEET in workbook.sheetnames:
        _validate_material_sheet(workbook[MATERIAL_SHEET], row_counts[MATERIAL_SHEET], errors)
    if LABOR_SHEET in workbook.sheetnames:
        _validate_labor_sheet(workbook[LABOR_SHEET], row_counts[LABOR_SHEET], errors)
    if EXPENSE_SHEET in workbook.sheetnames:
        _validate_expense_sheet(workbook[EXPENSE_SHEET], row_counts[EXPENSE_SHEET], errors)

    if errors:
        raise BudgetExcelValidationError(errors)

    return row_counts


def _parse_number_cell(value: Any) -> tuple[float, bool]:
    if value in (None, ""):
        return 0.0, True
    if isinstance(value, (int, float)):
        return float(value), True
    try:
        return float(str(value).strip().replace(",", "")), True
    except Exception:  # noqa: BLE001
        return 0.0, False


def _parse_material_execution_rows(ws, row_count: int, errors: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for offset in range(row_count):
        row_idx = DATA_START_ROW + offset
        equipment = _normalize_text(ws[f"B{row_idx}"].value)
        phase_raw = ws[f"C{row_idx}"].value
        unit_name = _normalize_text(ws[f"D{row_idx}"].value)
        part_name = _normalize_text(ws[f"E{row_idx}"].value)
        spec = _normalize_text(ws[f"F{row_idx}"].value)
        memo = _normalize_text(ws[f"K{row_idx}"].value)

        execution_amount, is_valid_amount = _parse_number_cell(ws[f"J{row_idx}"].value)
        if not is_valid_amount:
            _add_error(errors, f"재료비!J{row_idx}: 집행금액은 숫자여야 합니다.")
            continue

        phase = _normalize_phase_code(phase_raw)
        if not phase and (equipment or unit_name or part_name or spec or execution_amount or memo):
            _add_error(errors, f"재료비!C{row_idx}: 단계 값은 제작/설치만 허용됩니다.")
            continue

        has_identity = bool(equipment or unit_name or part_name or spec)
        has_input = abs(execution_amount) > 0 or bool(memo)

        if not has_identity and not has_input:
            continue
        if has_input and not equipment:
            _add_error(errors, f"재료비!B{row_idx}: 집행금액 입력 시 설비명이 필요합니다.")
            continue
        if not has_input:
            continue

        rows.append(
            {
                "equipment_name": equipment,
                "unit_name": unit_name,
                "part_name": part_name,
                "spec": spec,
                "executed_amount": execution_amount,
                "phase": phase,
                "memo": memo,
            }
        )
    return rows


def _parse_labor_execution_rows(ws, row_count: int, errors: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for offset in range(row_count):
        row_idx = DATA_START_ROW + offset
        equipment = _normalize_text(ws[f"B{row_idx}"].value)
        phase_raw = ws[f"C{row_idx}"].value
        staffing_type = _normalize_text(ws[f"D{row_idx}"].value) or "자체"
        task_name = _normalize_text(ws[f"E{row_idx}"].value)
        worker_type = _normalize_text(ws[f"F{row_idx}"].value)
        memo = _normalize_text(ws[f"L{row_idx}"].value)

        execution_amount, is_valid_amount = _parse_number_cell(ws[f"K{row_idx}"].value)
        if not is_valid_amount:
            _add_error(errors, f"인건비!K{row_idx}: 집행금액은 숫자여야 합니다.")
            continue

        phase = _normalize_phase_code(phase_raw)
        if not phase and (equipment or staffing_type or task_name or worker_type or execution_amount or memo):
            _add_error(errors, f"인건비!C{row_idx}: 단계 값은 제작/설치만 허용됩니다.")
            continue

        if staffing_type not in {"자체", "외주"}:
            _add_error(errors, f"인건비!D{row_idx}: 구분은 자체/외주만 허용됩니다.")
            continue

        has_identity = bool(equipment or task_name or worker_type)
        has_input = abs(execution_amount) > 0 or bool(memo)

        if not has_identity and not has_input:
            continue
        if has_input and not equipment:
            _add_error(errors, f"인건비!B{row_idx}: 집행금액 입력 시 설비명이 필요합니다.")
            continue
        if not has_input:
            continue

        rows.append(
            {
                "equipment_name": equipment,
                "task_name": task_name,
                "staffing_type": staffing_type,
                "worker_type": worker_type,
                "executed_amount": execution_amount,
                "phase": phase,
                "memo": memo,
            }
        )
    return rows


def _parse_expense_execution_rows(ws, row_count: int, errors: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for offset in range(row_count):
        row_idx = DATA_START_ROW + offset
        equipment = _normalize_text(ws[f"B{row_idx}"].value)
        phase_raw = ws[f"C{row_idx}"].value
        expense_type = _normalize_text(ws[f"D{row_idx}"].value) or "자체"
        expense_name = _normalize_text(ws[f"E{row_idx}"].value)
        basis = _normalize_text(ws[f"F{row_idx}"].value)
        memo = _normalize_text(ws[f"J{row_idx}"].value)

        execution_amount, is_valid_amount = _parse_number_cell(ws[f"I{row_idx}"].value)
        if not is_valid_amount:
            _add_error(errors, f"경비!I{row_idx}: 집행금액은 숫자여야 합니다.")
            continue

        phase = _normalize_phase_code(phase_raw)
        if not phase and (equipment or expense_type or expense_name or basis or execution_amount or memo):
            _add_error(errors, f"경비!C{row_idx}: 단계 값은 제작/설치만 허용됩니다.")
            continue

        if expense_type not in {"자체", "외주"}:
            _add_error(errors, f"경비!D{row_idx}: 구분은 자체/외주만 허용됩니다.")
            continue

        has_identity = bool(equipment or expense_name or basis)
        has_input = abs(execution_amount) > 0 or bool(memo)

        if not has_identity and not has_input:
            continue
        if has_input and not equipment:
            _add_error(errors, f"경비!B{row_idx}: 집행금액 입력 시 설비명이 필요합니다.")
            continue
        if not has_input:
            continue

        rows.append(
            {
                "equipment_name": equipment,
                "expense_type": expense_type,
                "expense_name": expense_name,
                "basis": basis,
                "executed_amount": execution_amount,
                "phase": phase,
                "memo": memo,
            }
        )
    return rows


def parse_budget_excel_execution_import(file_bytes: bytes) -> dict[str, Any]:
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=False)
    except Exception as exc:  # noqa: BLE001
        raise BudgetExcelValidationError([f"엑셀 파일을 열 수 없습니다: {exc}"]) from exc

    row_counts = _validate_template(workbook)

    errors: list[str] = []
    material_rows = _parse_material_execution_rows(workbook[MATERIAL_SHEET], row_counts[MATERIAL_SHEET], errors)
    labor_rows = _parse_labor_execution_rows(workbook[LABOR_SHEET], row_counts[LABOR_SHEET], errors)
    expense_rows = _parse_expense_execution_rows(workbook[EXPENSE_SHEET], row_counts[EXPENSE_SHEET], errors)

    if errors:
        raise BudgetExcelValidationError(errors)

    return {
        "execution_material_items": material_rows,
        "execution_labor_items": labor_rows,
        "execution_expense_items": expense_rows,
        "updated_counts": {
            "material": len(material_rows),
            "labor": len(labor_rows),
            "expense": len(expense_rows),
        },
    }
