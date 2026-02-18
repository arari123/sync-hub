import unittest
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.core.budget_excel import (
    BudgetExcelValidationError,
    EXPENSE_SHEET,
    LABOR_SHEET,
    MATERIAL_SHEET,
    META_SHEET,
    SUMMARY_SHEET,
    build_budget_excel_bytes,
    parse_budget_excel_execution_import,
)


class BudgetExcelTests(unittest.TestCase):
    def setUp(self):
        self.project = SimpleNamespace(id=11, code="SYNC-11", name="테스트 프로젝트")
        self.version = SimpleNamespace(id=101, version_no=2, revision_no=1)
        self.detail_payload = {
            "material_items": [
                {
                    "equipment_name": "설비-A",
                    "unit_name": "유닛-1",
                    "part_name": "파츠-1",
                    "spec": "M3",
                    "quantity": 5,
                    "unit_price": 1000,
                    "phase": "fabrication",
                    "memo": "",
                }
            ],
            "labor_items": [
                {
                    "equipment_name": "설비-A",
                    "task_name": "조립",
                    "staffing_type": "자체",
                    "worker_type": "PM",
                    "unit": "H",
                    "quantity": 8,
                    "headcount": 1,
                    "phase": "fabrication",
                    "memo": "",
                }
            ],
            "expense_items": [
                {
                    "equipment_name": "설비-A",
                    "expense_type": "자체",
                    "expense_name": "출장비",
                    "basis": "기본",
                    "quantity": 1,
                    "amount": 100000,
                    "phase": "fabrication",
                    "memo": "",
                }
            ],
            "execution_material_items": [],
            "execution_labor_items": [],
            "execution_expense_items": [],
            "budget_settings": {
                "installation_locale": "domestic",
                "labor_days_per_week_domestic": 5,
                "labor_days_per_week_overseas": 7,
                "labor_days_per_month_domestic": 22,
                "labor_days_per_month_overseas": 30,
            },
        }

    def _build_bytes(self):
        return build_budget_excel_bytes(
            project=self.project,
            version=self.version,
            detail_payload=self.detail_payload,
        )

    @staticmethod
    def _save_workbook(workbook):
        buf = BytesIO()
        workbook.save(buf)
        return buf.getvalue()

    def test_build_excel_contains_required_sheets(self):
        raw = self._build_bytes()
        workbook = load_workbook(BytesIO(raw), data_only=False)
        self.assertEqual(workbook.sheetnames, [SUMMARY_SHEET, MATERIAL_SHEET, LABOR_SHEET, EXPENSE_SHEET, META_SHEET])

    def test_parse_execution_import_success(self):
        workbook = load_workbook(BytesIO(self._build_bytes()), data_only=False)

        workbook[MATERIAL_SHEET]["J5"] = 1111
        workbook[MATERIAL_SHEET]["K5"] = "재료 집행"
        workbook[LABOR_SHEET]["K5"] = 2222
        workbook[LABOR_SHEET]["L5"] = "인건비 집행"
        workbook[EXPENSE_SHEET]["I5"] = 3333
        workbook[EXPENSE_SHEET]["J5"] = "경비 집행"

        parsed = parse_budget_excel_execution_import(self._save_workbook(workbook))

        self.assertEqual(parsed["updated_counts"]["material"], 1)
        self.assertEqual(parsed["updated_counts"]["labor"], 1)
        self.assertEqual(parsed["updated_counts"]["expense"], 1)

        self.assertEqual(parsed["execution_material_items"][0]["executed_amount"], 1111)
        self.assertEqual(parsed["execution_labor_items"][0]["executed_amount"], 2222)
        self.assertEqual(parsed["execution_expense_items"][0]["executed_amount"], 3333)

    def test_parse_fails_when_header_changed(self):
        workbook = load_workbook(BytesIO(self._build_bytes()), data_only=False)
        workbook[MATERIAL_SHEET]["B4"] = "설비"

        with self.assertRaises(BudgetExcelValidationError) as raised:
            parse_budget_excel_execution_import(self._save_workbook(workbook))

        self.assertIn("재료비!B4", str(raised.exception))

    def test_parse_fails_when_formula_removed(self):
        workbook = load_workbook(BytesIO(self._build_bytes()), data_only=False)
        workbook[MATERIAL_SHEET]["I5"] = 0

        with self.assertRaises(BudgetExcelValidationError) as raised:
            parse_budget_excel_execution_import(self._save_workbook(workbook))

        self.assertIn("재료비!I5", str(raised.exception))

    def test_parse_fails_when_sheet_protection_disabled(self):
        workbook = load_workbook(BytesIO(self._build_bytes()), data_only=False)
        workbook[MATERIAL_SHEET].protection.sheet = False

        with self.assertRaises(BudgetExcelValidationError) as raised:
            parse_budget_excel_execution_import(self._save_workbook(workbook))

        self.assertIn("시트 보호", str(raised.exception))


if __name__ == "__main__":
    unittest.main()
