#!/usr/bin/env python3
"""Reset budget/demo data and seed mock projects with spreadsheet documents."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import text

from app import models
from app.core.auth_utils import to_iso, utcnow
from app.core.budget_logic import aggregate_equipment_costs_from_detail, detail_payload_to_json
from app.database import SessionLocal, ensure_runtime_schema

UPLOAD_DIR = Path('uploads')


def _build_detail_payload(project_name: str, execution_ratio: float) -> dict:
    equip_a = f"{project_name}-설비A"
    equip_b = f"{project_name}-설비B"

    material_items = [
        {
            'equipment_name': equip_a,
            'unit_name': 'FEED',
            'part_name': '리니어 모듈',
            'spec': 'LM-220',
            'quantity': 2,
            'unit_price': 1800000,
            'phase': 'fabrication',
            'memo': '핵심 구동 파츠',
        },
        {
            'equipment_name': equip_a,
            'unit_name': 'VISION',
            'part_name': '산업용 카메라',
            'spec': 'VC-12MP',
            'quantity': 3,
            'unit_price': 950000,
            'phase': 'installation',
            'memo': '정렬 캘리브레이션 포함',
        },
        {
            'equipment_name': equip_b,
            'unit_name': 'ROBOT',
            'part_name': '서보 드라이브',
            'spec': 'SD-3KW',
            'quantity': 4,
            'unit_price': 640000,
            'phase': 'fabrication',
            'memo': '축 제어용',
        },
        {
            'equipment_name': equip_b,
            'unit_name': 'PANEL',
            'part_name': 'I/O 모듈',
            'spec': 'IO-128',
            'quantity': 2,
            'unit_price': 520000,
            'phase': 'installation',
            'memo': '현장 배선 포함',
        },
    ]

    labor_items = [
        {
            'equipment_name': equip_a,
            'task_name': '프레임 조립',
            'worker_type': '기계조립',
            'unit': 'H',
            'quantity': 56,
            'hourly_rate': 42000,
            'phase': 'fabrication',
            'memo': '2인 3.5일 기준',
        },
        {
            'equipment_name': equip_a,
            'task_name': '시운전/튜닝',
            'worker_type': '제어엔지니어',
            'unit': 'H',
            'quantity': 32,
            'hourly_rate': 65000,
            'phase': 'installation',
            'memo': '현장 설치',
        },
        {
            'equipment_name': equip_b,
            'task_name': '배선/패널 작업',
            'worker_type': '전장기사',
            'unit': 'H',
            'quantity': 48,
            'hourly_rate': 46000,
            'phase': 'fabrication',
            'memo': '제작동 작업',
        },
        {
            'equipment_name': equip_b,
            'task_name': '현장 인터락 점검',
            'worker_type': '안전검증',
            'unit': 'H',
            'quantity': 20,
            'hourly_rate': 59000,
            'phase': 'installation',
            'memo': '설치 후 검수',
        },
    ]

    expense_items = [
        {
            'equipment_name': equip_a,
            'expense_name': '가공/외주비',
            'basis': '프레임 CNC',
            'amount': 2300000,
            'phase': 'fabrication',
            'memo': '외주 가공처 1곳',
        },
        {
            'equipment_name': equip_a,
            'expense_name': '현장 운송비',
            'basis': '5톤 트럭 1회',
            'amount': 640000,
            'phase': 'installation',
            'memo': '왕복 기준',
        },
        {
            'equipment_name': equip_b,
            'expense_name': '보호구/소모품',
            'basis': '안전장구+체결류',
            'amount': 420000,
            'phase': 'fabrication',
            'memo': '월간 패키지',
        },
        {
            'equipment_name': equip_b,
            'expense_name': '설치 장비 렌탈',
            'basis': '리프트 2일',
            'amount': 510000,
            'phase': 'installation',
            'memo': '현장 반입',
        },
    ]

    execution_material_items = [
        {
            'equipment_name': equip_a,
            'unit_name': 'FEED-실집행',
            'part_name': '리니어 블록',
            'spec': 'LM-220A',
            'executed_amount': round(3250000 * execution_ratio),
            'phase': 'fabrication',
            'memo': '예산 파츠와 다르게 집행',
        },
        {
            'equipment_name': equip_a,
            'unit_name': 'VISION-현장',
            'part_name': '카메라 브라켓',
            'spec': 'BRK-12',
            'executed_amount': round(860000 * execution_ratio),
            'phase': 'installation',
            'memo': '현장 추가 구매',
        },
        {
            'equipment_name': equip_b,
            'unit_name': 'ROBOT-실집행',
            'part_name': '감속기 세트',
            'spec': 'RG-90',
            'executed_amount': round(2410000 * execution_ratio),
            'phase': 'fabrication',
            'memo': '유닛/파츠 구조 변경',
        },
        {
            'equipment_name': equip_b,
            'unit_name': 'SITE-ADAPTER',
            'part_name': '배선 어댑터',
            'spec': 'ADP-24',
            'executed_amount': round(490000 * execution_ratio),
            'phase': 'installation',
            'memo': '설치 대응 부품',
        },
    ]

    execution_labor_items = [
        {
            'equipment_name': equip_a,
            'task_name': '사전 셋업',
            'worker_type': '현장기술',
            'executed_amount': round(2100000 * execution_ratio),
            'phase': 'fabrication',
            'memo': '예산 작업명과 상이',
        },
        {
            'equipment_name': equip_a,
            'task_name': '현장 안정화',
            'worker_type': '제어튜닝',
            'executed_amount': round(1680000 * execution_ratio),
            'phase': 'installation',
            'memo': '추가 튜닝 반영',
        },
        {
            'equipment_name': equip_b,
            'task_name': '배관/배선 변경',
            'worker_type': '전장협력사',
            'executed_amount': round(1940000 * execution_ratio),
            'phase': 'fabrication',
            'memo': '협력사 투입',
        },
        {
            'equipment_name': equip_b,
            'task_name': '인터락 재검증',
            'worker_type': '안전점검',
            'executed_amount': round(1080000 * execution_ratio),
            'phase': 'installation',
            'memo': '재검증 2회',
        },
    ]

    execution_expense_items = [
        {
            'equipment_name': equip_a,
            'expense_name': '긴급 가공비',
            'basis': '변경도면 반영',
            'executed_amount': round(1980000 * execution_ratio),
            'phase': 'fabrication',
            'memo': '예산 대비 구조 변경',
        },
        {
            'equipment_name': equip_a,
            'expense_name': '추가 물류비',
            'basis': '야간 반입',
            'executed_amount': round(520000 * execution_ratio),
            'phase': 'installation',
            'memo': '야간작업',
        },
        {
            'equipment_name': equip_b,
            'expense_name': '외주 공임비',
            'basis': '현장 개조',
            'executed_amount': round(640000 * execution_ratio),
            'phase': 'fabrication',
            'memo': '개조 작업 반영',
        },
        {
            'equipment_name': equip_b,
            'expense_name': '장비 대여비',
            'basis': '리프트 추가 1일',
            'executed_amount': round(430000 * execution_ratio),
            'phase': 'installation',
            'memo': '추가 일수 발생',
        },
    ]

    return {
        'material_items': material_items,
        'labor_items': labor_items,
        'expense_items': expense_items,
        'execution_material_items': execution_material_items,
        'execution_labor_items': execution_labor_items,
        'execution_expense_items': execution_expense_items,
    }


def _write_standard_excel(path: Path, detail: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'BudgetInput'
    ws.append(['구분', '설비', '유닛/작업', '항목', '규격/기준', '수량', '단가/금액', '단계', '집행'])

    for item in detail['material_items']:
        ws.append([
            '재료비',
            item['equipment_name'],
            item['unit_name'],
            item['part_name'],
            item['spec'],
            item['quantity'],
            item['unit_price'],
            '제작' if item['phase'] == 'fabrication' else '설치',
            '',
        ])

    for item in detail['labor_items']:
        ws.append([
            '인건비',
            item['equipment_name'],
            item['worker_type'],
            item['task_name'],
            item['unit'],
            item['quantity'],
            item['hourly_rate'],
            '제작' if item['phase'] == 'fabrication' else '설치',
            '',
        ])

    for item in detail['expense_items']:
        ws.append([
            '경비',
            item['equipment_name'],
            '-',
            item['expense_name'],
            item['basis'],
            1,
            item['amount'],
            '제작' if item['phase'] == 'fabrication' else '설치',
            '',
        ])

    for item in detail['execution_material_items']:
        ws.append([
            '재료비-집행',
            item['equipment_name'],
            item['unit_name'],
            item['part_name'],
            item['spec'],
            '',
            '',
            '제작' if item['phase'] == 'fabrication' else '설치',
            item['executed_amount'],
        ])

    for item in detail['execution_labor_items']:
        ws.append([
            '인건비-집행',
            item['equipment_name'],
            item['worker_type'],
            item['task_name'],
            '-',
            '',
            '',
            '제작' if item['phase'] == 'fabrication' else '설치',
            item['executed_amount'],
        ])

    for item in detail['execution_expense_items']:
        ws.append([
            '경비-집행',
            item['equipment_name'],
            '-',
            item['expense_name'],
            item['basis'],
            '',
            '',
            '제작' if item['phase'] == 'fabrication' else '설치',
            item['executed_amount'],
        ])

    wb.save(path)


def _write_multisheet_excel(path: Path, detail: dict) -> None:
    wb = Workbook()
    ws_material = wb.active
    ws_material.title = 'Material'
    ws_material.append(['구분', '설비', '유닛', '부품/작업', '규격', '수량', '단가', '단계', '집행'])
    for item in detail['material_items']:
        ws_material.append([
            'budget',
            item['equipment_name'],
            item['unit_name'],
            item['part_name'],
            item['spec'],
            item['quantity'],
            item['unit_price'],
            item['phase'],
            '',
        ])
    for item in detail['execution_material_items']:
        ws_material.append([
            'execution',
            item['equipment_name'],
            item['unit_name'],
            item['part_name'],
            item['spec'],
            '',
            '',
            item['phase'],
            item['executed_amount'],
        ])

    ws_labor = wb.create_sheet('Labor')
    ws_labor.append(['구분', '설비', '작업명', '직군', '단위', '수량', '단가', '단계', '집행'])
    for item in detail['labor_items']:
        ws_labor.append([
            'budget',
            item['equipment_name'],
            item['task_name'],
            item['worker_type'],
            item['unit'],
            item['quantity'],
            item['hourly_rate'],
            item['phase'],
            '',
        ])
    for item in detail['execution_labor_items']:
        ws_labor.append([
            'execution',
            item['equipment_name'],
            item['task_name'],
            item['worker_type'],
            '-',
            '',
            '',
            item['phase'],
            item['executed_amount'],
        ])

    ws_expense = wb.create_sheet('Expense')
    ws_expense.append(['구분', '설비', '경비명', '산정기준', '예산금액', '단계', '집행'])
    for item in detail['expense_items']:
        ws_expense.append([
            'budget',
            item['equipment_name'],
            item['expense_name'],
            item['basis'],
            item['amount'],
            item['phase'],
            '',
        ])
    for item in detail['execution_expense_items']:
        ws_expense.append([
            'execution',
            item['equipment_name'],
            item['expense_name'],
            item['basis'],
            '',
            item['phase'],
            item['executed_amount'],
        ])

    wb.save(path)


def _write_merged_header_excel(path: Path, detail: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'ExecutionLayout'
    ws.merge_cells('A1:C1')
    ws.merge_cells('D1:F1')
    ws['A1'] = '기본 정보'
    ws['D1'] = '금액 정보'
    ws.append(['설비', '항목', '단계', '예산', '집행', '비고'])

    for item in detail['execution_material_items'][:2]:
        ws.append([
            item['equipment_name'],
            item['part_name'],
            item['phase'],
            '',
            item['executed_amount'],
            item['memo'],
        ])

    for item in detail['execution_labor_items'][:2]:
        ws.append([
            item['equipment_name'],
            item['task_name'],
            item['phase'],
            '',
            item['executed_amount'],
            item['memo'],
        ])

    wb.save(path)


def _write_csv(path: Path, detail: dict) -> None:
    with path.open('w', encoding='utf-8', newline='') as fp:
        writer = csv.writer(fp)
        writer.writerow(['category', 'row_type', 'equipment', 'name', 'phase', 'budget', 'executed'])
        for item in detail['material_items']:
            writer.writerow([
                'material',
                'budget',
                item['equipment_name'],
                item['part_name'],
                item['phase'],
                item['quantity'] * item['unit_price'],
                '',
            ])
        for item in detail['execution_material_items']:
            writer.writerow([
                'material',
                'execution',
                item['equipment_name'],
                item['part_name'],
                item['phase'],
                '',
                item['executed_amount'],
            ])
        for item in detail['labor_items']:
            writer.writerow([
                'labor',
                'budget',
                item['equipment_name'],
                item['task_name'],
                item['phase'],
                item['quantity'] * item['hourly_rate'],
                '',
            ])
        for item in detail['execution_labor_items']:
            writer.writerow([
                'labor',
                'execution',
                item['equipment_name'],
                item['task_name'],
                item['phase'],
                '',
                item['executed_amount'],
            ])
        for item in detail['expense_items']:
            writer.writerow([
                'expense',
                'budget',
                item['equipment_name'],
                item['expense_name'],
                item['phase'],
                item['amount'],
                '',
            ])
        for item in detail['execution_expense_items']:
            writer.writerow([
                'expense',
                'execution',
                item['equipment_name'],
                item['expense_name'],
                item['phase'],
                '',
                item['executed_amount'],
            ])


def main() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    ensure_runtime_schema()
    session = SessionLocal()
    try:
        active_users = (
            session.query(models.User)
            .filter(models.User.is_active.is_(True), models.User.email_verified.is_(True))
            .order_by(models.User.id.asc())
            .all()
        )
        if not active_users:
            raise RuntimeError('활성/인증된 사용자 계정을 찾을 수 없습니다.')

        now_iso = to_iso(utcnow())

        session.execute(text('TRUNCATE TABLE budget_equipments, budget_versions, budget_projects RESTART IDENTITY CASCADE;'))
        session.execute(text('TRUNCATE TABLE dedup_audit_log, dedup_cluster_members, dedup_clusters, documents RESTART IDENTITY CASCADE;'))
        session.commit()

        for ext in ('*.xlsx', '*.xlsm', '*.xltx', '*.xltm', '*.csv'):
            for file_path in UPLOAD_DIR.glob(ext):
                file_path.unlink(missing_ok=True)

        seed_templates = [
            {
                'name': '반도체 검사라인 A 개선',
                'code': 'MOCK-REV-001',
                'description': '검토 단계 예산 산정용 가상 프로젝트',
                'project_type': 'equipment',
                'customer_name': '가상반도체',
                'installation_site': '화성 1공장',
                'stage': 'review',
                'status': 'confirmed',
                'execution_ratio': 0.0,
            },
            {
                'name': '2차전지 조립 셀 제작',
                'code': 'MOCK-FAB-002',
                'description': '제작 단계 집행금액 입력 테스트 프로젝트',
                'project_type': 'equipment',
                'customer_name': '가상배터리',
                'installation_site': '오창 파일럿동',
                'stage': 'fabrication',
                'status': 'confirmed',
                'execution_ratio': 0.52,
            },
            {
                'name': '물류센터 AGV 설치',
                'code': 'MOCK-INS-003',
                'description': '설치 단계 집행금액 입력 테스트 프로젝트',
                'project_type': 'parts',
                'customer_name': '가상물류',
                'installation_site': '평택 허브센터',
                'stage': 'installation',
                'status': 'confirmed',
                'execution_ratio': 0.74,
            },
            {
                'name': '레이저 마커 유지보수 AS',
                'code': 'MOCK-AS-004',
                'description': 'AS 단계 집행관리 테스트 프로젝트',
                'project_type': 'as',
                'customer_name': '가상전자',
                'installation_site': '구미 생산기술팀',
                'stage': 'warranty',
                'status': 'confirmed',
                'execution_ratio': 0.91,
            },
        ]

        seed_projects = []
        for index, user in enumerate(active_users):
            template = seed_templates[index % len(seed_templates)]
            email_prefix = (user.email or f'user{user.id}').split('@', 1)[0][:12]
            seed_projects.append(
                {
                    **template,
                    'name': f"{template['name']} ({email_prefix})",
                    'code': f"{template['code']}-U{int(user.id):02d}",
                    'manager_user_id': int(user.id),
                    'created_by_user_id': int(user.id),
                }
            )

        created_projects: list[tuple[models.BudgetProject, dict]] = []
        for item in seed_projects:
            project = models.BudgetProject(
                name=item['name'],
                code=item['code'],
                description=item['description'],
                project_type=item['project_type'],
                customer_name=item['customer_name'],
                installation_site=item['installation_site'],
                created_by_user_id=int(item['created_by_user_id']),
                manager_user_id=int(item['manager_user_id']),
                current_stage=item['stage'],
                created_at=now_iso,
                updated_at=now_iso,
            )
            session.add(project)
            session.flush()

            detail_payload = _build_detail_payload(item['name'], float(item['execution_ratio']))
            version = models.BudgetVersion(
                project_id=project.id,
                stage=item['stage'],
                status=item['status'],
                version_no=1,
                revision_no=0,
                parent_version_id=None,
                change_reason='',
                budget_detail_json=detail_payload_to_json(detail_payload),
                is_current=True,
                confirmed_at=now_iso if item['status'] == 'confirmed' else None,
                created_at=now_iso,
                updated_at=now_iso,
            )
            session.add(version)
            session.flush()

            aggregates = aggregate_equipment_costs_from_detail(detail_payload)
            for index, aggregate in enumerate(aggregates):
                session.add(
                    models.BudgetEquipment(
                        version_id=version.id,
                        equipment_name=aggregate['equipment_name'],
                        material_fab_cost=aggregate['material_fab_cost'],
                        material_install_cost=aggregate['material_install_cost'],
                        labor_fab_cost=aggregate['labor_fab_cost'],
                        labor_install_cost=aggregate['labor_install_cost'],
                        expense_fab_cost=aggregate['expense_fab_cost'],
                        expense_install_cost=aggregate['expense_install_cost'],
                        currency='KRW',
                        sort_order=index,
                        created_at=now_iso,
                        updated_at=now_iso,
                    )
                )

            created_projects.append((project, detail_payload))

        session.flush()

        spreadsheet_specs = [
            ('mock_budget_matrix.xlsx', _write_standard_excel),
            ('mock_budget_multisheet.xlsx', _write_multisheet_excel),
            ('mock_budget_merged_header.xlsx', _write_merged_header_excel),
            ('mock_budget_snapshot.csv', _write_csv),
        ]

        for index, (filename, writer) in enumerate(spreadsheet_specs):
            project, detail_payload = created_projects[index % len(created_projects)]
            file_path = UPLOAD_DIR / filename
            writer(file_path, detail_payload)

            session.add(
                models.Document(
                    filename=filename,
                    file_path=str(file_path),
                    status='completed',
                    content_text=f"{project.name} 예산/집행 가상 문서",
                    document_types='spreadsheet,budget',
                    ai_title=f"{project.name} 가상 엑셀",
                    ai_summary_short='가상 프로젝트 테스트용 엑셀 문서',
                    created_at=now_iso,
                    project_id=project.id,
                    dedup_status='unique',
                )
            )

        session.commit()

        project_count = session.query(models.BudgetProject).count()
        version_count = session.query(models.BudgetVersion).count()
        document_count = session.query(models.Document).count()

        manager_ids = sorted({int(item['manager_user_id']) for item in seed_projects})
        print(f'[seed] manager_user_ids={manager_ids}')
        print(f'[seed] projects={project_count}, versions={version_count}, documents={document_count}')
        print('[seed] files:', ', '.join(name for name, _ in spreadsheet_specs))
    finally:
        session.close()


if __name__ == '__main__':
    main()
