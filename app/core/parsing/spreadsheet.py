from __future__ import annotations

import csv
import os
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

from ..chunking.chunker import SourceSegment, table_group_to_structured_text
from .cleaning import normalize_line, normalize_text

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

_SUPPORTED_SPREADSHEET_EXTENSIONS = {
    ".xlsx",
    ".xlsm",
    ".xltx",
    ".xltm",
    ".csv",
}


def is_spreadsheet_file(file_path: str) -> bool:
    suffix = Path(file_path).suffix.lower()
    return suffix in _SUPPORTED_SPREADSHEET_EXTENSIONS


def _normalize_cell(value) -> str:  # noqa: ANN001
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")

    if isinstance(value, float):
        text = f"{value:.6f}".rstrip("0").rstrip(".")
        return text or "0"

    return normalize_line(str(value).replace("\n", " ").strip())


def _build_sheet_segments(
    sheet_name: str,
    sheet_index: int,
    rows: Iterable[Sequence[object]],
) -> Tuple[List[SourceSegment], List[str], List[str]]:
    key_value_lines: List[str] = []
    table_lines: List[str] = []

    for row in rows:
        cells = [_normalize_cell(cell) for cell in row]
        cells = [cell for cell in cells if cell]
        if not cells:
            continue

        if len(cells) == 2:
            left, right = cells
            key_value_lines.append(f"{left}: {right}")

        table_lines.append(" | ".join(cells))

    if not key_value_lines and not table_lines:
        return [], [], []

    segments: List[SourceSegment] = []
    raw_parts: List[str] = []
    clean_parts: List[str] = []

    if key_value_lines:
        paragraph_text = normalize_text("\n".join(key_value_lines))
        if paragraph_text:
            segments.append(
                SourceSegment(
                    page=sheet_index,
                    chunk_type="paragraph",
                    text=paragraph_text,
                    raw_text="\n".join(key_value_lines).strip(),
                    section_title=sheet_name,
                )
            )
            raw_parts.append(paragraph_text)
            clean_parts.append(paragraph_text)

    if table_lines:
        table_raw, row_sentences = table_group_to_structured_text(table_lines)
        table_raw = normalize_text(table_raw)
        raw_text = "\n".join(table_lines).strip()

        if table_raw:
            segments.append(
                SourceSegment(
                    page=sheet_index,
                    chunk_type="table_raw",
                    text=table_raw,
                    raw_text=raw_text,
                    section_title=sheet_name,
                )
            )
            raw_parts.append(raw_text)
            clean_parts.append(table_raw)

        for row_sentence in row_sentences:
            cleaned = normalize_text(row_sentence)
            if not cleaned:
                continue
            segments.append(
                SourceSegment(
                    page=sheet_index,
                    chunk_type="table_row_sentence",
                    text=cleaned,
                    raw_text=raw_text,
                    section_title=sheet_name,
                )
            )

    return segments, raw_parts, clean_parts


def _extract_from_csv(file_path: str) -> Tuple[str, str, List[SourceSegment]]:
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = [row for row in reader]

    segments, raw_parts, clean_parts = _build_sheet_segments(
        sheet_name="CSV",
        sheet_index=1,
        rows=rows,
    )

    raw_text = normalize_text("\n\n".join(raw_parts))
    clean_text = normalize_text("\n\n".join(clean_parts))
    return raw_text, clean_text, segments


def _extract_from_xlsx(file_path: str) -> Tuple[str, str, List[SourceSegment]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed.")

    workbook = load_workbook(file_path, read_only=True, data_only=True)

    all_segments: List[SourceSegment] = []
    raw_parts: List[str] = []
    clean_parts: List[str] = []

    for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
        worksheet = workbook[sheet_name]
        sheet_rows = worksheet.iter_rows(values_only=True)

        segments, sheet_raw_parts, sheet_clean_parts = _build_sheet_segments(
            sheet_name=sheet_name,
            sheet_index=sheet_index,
            rows=sheet_rows,
        )

        if segments:
            all_segments.extend(segments)
            raw_parts.append(f"[{sheet_name}]")
            raw_parts.extend(sheet_raw_parts)
            clean_parts.extend(sheet_clean_parts)

    workbook.close()

    raw_text = normalize_text("\n\n".join(raw_parts))
    clean_text = normalize_text("\n\n".join(clean_parts))
    return raw_text, clean_text, all_segments


def extract_spreadsheet_segments(file_path: str) -> Tuple[str, str, List[SourceSegment]]:
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return _extract_from_csv(file_path)

    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _extract_from_xlsx(file_path)

    return "", "", []
